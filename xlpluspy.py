from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Создание объекта
wb = Workbook()


# Загружаем эксель файл
wb = load_workbook('123.xlsx')

# Открываем лист (первый)
ws = wb.active


# Запись в переменную зачение ячейки
model = ws['A5'].value
color = ws['B5'].value
availability = ws['E5'].value
price = ws['C5'].value

# Вывод записанных переменных
print(f'Iphone {model}: {color}, Стоимость: {price}. Наличие: {availability}.')

# Запись строки в переменную
row_2 = ws[2]

print(row_2)
# Вывод этой строки
for cell in row_2:
    print(f'{cell.value} (row)')

# Значения из диапазона
range = ws['A2':'D10']

# Вывод значений в ячейках из указанного диапазона
print('+++')
for cell in range:
    print('+++\n')
    for x in cell:
        print(x.value)

# Вывод моделей, которых НЕТ в наличии:
not_Av = ws['A2':'E10']
for i in not_Av:
   k = 0
   if i[4].value == 'Нет':
      print('---\n')
      print('НЕТ В НАЛИЧИИ (Iphone):')
      for j in i:
        k += 1
        if k < 3:
            print(j.value)
        elif k == 4:
           print(j.value)